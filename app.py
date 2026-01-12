from flask import (
    Flask, render_template, request,
    redirect, url_for, session,
    flash, jsonify
)
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import time
import json
import os
import re
from openpyxl import load_workbook

def slugify(text):
    text = text.lower()
    text = re.sub(r"[àáạảãâầấậẩẫăằắặẳẵ]", "a", text)
    text = re.sub(r"[èéẹẻẽêềếệểễ]", "e", text)
    text = re.sub(r"[ìíịỉĩ]", "i", text)
    text = re.sub(r"[òóọỏõôồốộổỗơờớợởỡ]", "o", text)
    text = re.sub(r"[ùúụủũưừứựửữ]", "u", text)
    text = re.sub(r"[ỳýỵỷỹ]", "y", text)
    text = re.sub(r"đ", "d", text)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")

#=========tạo quizid tự động========
def make_quiz_id(subject, class_name, quiz_name):
    # subject: hoa -> h
    s = subject[0].lower()

    # class: lop 8 / hsk 1 -> 8 / 1
    nums = re.findall(r"\d+", class_name)
    c = nums[0] if nums else "0"

    # quiz_name: Hóa học Cơ bản -> hhcb
    words = re.findall(r"[A-Za-zÀ-ỹ]+", quiz_name)
    n = "".join(w[0].lower() for w in words)

    return f"{s}{c}-{n}"


from functools import wraps
#==========KIEM TRA QUYEN ADMIN=======
from functools import wraps
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            return redirect(url_for("choose_subject"))
        return f(*args, **kwargs)
    return wrapper



app = Flask(__name__)
app.secret_key = "quiz_secret_key"
SUBJECT_SHEET_MAP = {
    "list": "LIST",
    "toan": "TOAN",
    "ly": "LY",
    "hoa": "HOA",
    "trung": "CHINA",
}

# ================= GOOGLE SHEET =================
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

service_account_info = json.loads(
    os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "{}")
)

creds = Credentials.from_service_account_info(
    service_account_info,
    scopes=SCOPES
)

gc = gspread.authorize(creds)


SHEET_ID = "16Bnz7F28gs4Fj3bH0XjRpQ-_4jcIWCbh37XptYecesU"
sh = gc.open_by_key(SHEET_ID)

# ================= CACHE =================
QUIZ_CACHE = {}
CACHE_TIME = 0
CACHE_EXPIRE = 300  # 5 phút


def load_quiz_from_sheet():
    global QUIZ_CACHE, CACHE_TIME

    #print("Load quiz từ Google Sheet")
    QUIZ_CACHE = {}

    for key, sheet_name in SUBJECT_SHEET_MAP.items():
        ws = sh.worksheet(sheet_name)
        QUIZ_CACHE[key] = ws.get_all_records()

    CACHE_TIME = time.time()



def get_quiz(key):
    if time.time() - CACHE_TIME > CACHE_EXPIRE or not QUIZ_CACHE:
        load_quiz_from_sheet()
    return QUIZ_CACHE.get(key, [])


# ================= USERS =================
ws_user = sh.worksheet("USERS")

# ================= LOGIN =================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Vui lòng nhập đầy đủ tài khoản và mật khẩu", "error")
            return render_template("login.html")

        rows = ws_user.get_all_values()[1:]
        for row in rows:
            if row[0] == username and row[1] == password:
                session["user"] = username
                session["role"] = row[4].strip().lower() if len(row) > 4 else "user"

                if session["role"] == "admin":
                    return redirect(url_for("admin"))
                else:
                    return redirect(url_for("choose_subject"))

        flash("Sai tài khoản hoặc mật khẩu", "error")

    return render_template("login.html")


# ================= REGISTER =================
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()
        fullname = request.form.get("fullname", "").strip()
        phone = request.form.get("phone", "").strip()

        if not all([username, password, confirm, fullname, phone]):
            flash("Vui lòng nhập đầy đủ thông tin", "error")
            return render_template("register.html")

        if password != confirm:
            flash("Mật khẩu xác nhận không khớp", "error")
            return render_template("register.html")

        rows = ws_user.get_all_values()[1:]
        for row in rows:
            if row and row[0] == username:
                flash("Tài khoản đã tồn tại", "error")
                return render_template("register.html")

        ws_user.append_row([
            username,
            password,
            fullname,
            phone,
            "user",  # 👈 role mặc định
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

        flash("Đăng ký thành công, mời đăng nhập", "success")
        return redirect(url_for("login"))

    return render_template("register.html")
#===========PAGE CHO ADMIN====================
# ================= PAGE ADMIN =================
@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin():
    # ======================================================
    # ===== 1. LOAD DANH SÁCH ĐỀ (LUÔN LOAD – GET & POST) ===
    # ======================================================
    ws_list = sh.worksheet("LIST")
    quizzes = ws_list.get_all_records()

    # ======================================================
    # ===== 2. LẤY ĐỀ ĐANG CHỌN (TỪ QUERY STRING – GET) =====
    # ======================================================
    selected_quiz = request.args.get("quiz_id")
    selected_subject = request.args.get("subject")

    questions = []

    # ======================================================
    # ===== 3. LOAD CÂU HỎI THEO ĐỀ ĐANG CHỌN ==============
    # ======================================================
    if selected_quiz and selected_subject:
        sheet_name = SUBJECT_SHEET_MAP.get(selected_subject)
        if sheet_name:
            ws = sh.worksheet(sheet_name)
            questions = [
                q for q in ws.get_all_records()
                if q.get("quiz_id") == selected_quiz
            ]

    # ======================================================
    # ===== 4. XỬ LÝ FORM POST =============================
    # ======================================================
    if request.method == "POST":
        action = request.form.get("action")

        # ==========================================
        # ===== 4.1 BULK UPLOAD (PASTE TEXT) =======
        # ==========================================
        if action == "bulk_upload":
            subject = request.form["subject"].strip().lower()
            class_name = request.form["class_name"].strip().lower()
            raw_text = request.form["bulk_text"]

            lines = [l for l in raw_text.splitlines() if l.strip()]
            if not lines:
                flash("❌ Không có dữ liệu", "error")
                return redirect("/admin")

            parsed = []
            for i, line in enumerate(lines, start=1):
                cols = line.split("\t")
                if len(cols) != 6:
                    flash(f"❌ Dòng {i} sai định dạng (cần 6 cột)", "error")
                    return redirect("/admin")

                parsed.append(cols)

            quiz_name = request.form["quiz_name"].strip()
            quiz_id = make_quiz_id(subject, class_name, quiz_name)

            # thêm vào LIST nếu chưa có
            if not any(q["quiz_id"] == quiz_id for q in quizzes):
                ws_list.append_row([subject, class_name, quiz_id, quiz_name])

            ws_list.append_row([
                subject,
                class_name,
                quiz_id,
                quiz_name,
                request.form.get("time_limit", 15)
            ])

            for row in parsed:
                question, a, b, c, d, correct = row

                ws_subject.append_row([
                    quiz_id,
                    quiz_name,
                    question.strip(),
                    a.strip(),
                    b.strip(),
                    c.strip(),
                    d.strip(),
                    correct.strip().upper()
                ])

            QUIZ_CACHE.clear()
            flash(f"✅ Đã upload {len(parsed)} câu hỏi", "success")
            return redirect(f"/admin?subject={subject}&quiz_id={quiz_id}")

        # ==========================================
        # ===== 4.2 UPLOAD FILE EXCEL ==============
        # ==========================================
        elif action == "upload_excel":
            subject = request.form["subject"].strip().lower()
            class_name = request.form["class_name"].strip().lower()
            file = request.files.get("excel_file")

            if not file:
                flash("❌ Không có file Excel", "error")
                return redirect("/admin")

            wb = load_workbook(file)
            ws_excel = wb.active
            rows = list(ws_excel.iter_rows(min_row=2, values_only=True))

            if not rows:
                flash("❌ File Excel rỗng", "error")
                return redirect("/admin")

            quiz_name = request.form["quiz_name"].strip()
            quiz_id = make_quiz_id(subject, class_name, quiz_name)

            if not any(q["quiz_id"] == quiz_id for q in quizzes):
                ws_list.append_row([
                    subject,
                    class_name,
                    quiz_id,
                    quiz_name,
                    request.form.get("time_limit", 15)
                ])

            ws_subject = sh.worksheet(SUBJECT_SHEET_MAP[subject])

            for r in rows:
                qname, question, a, b, c, d, correct = r
                ws_subject.append_row([
                    quiz_id,
                    str(qname).strip(),
                    str(question).strip(),
                    str(a).strip(),
                    str(b).strip(),
                    str(c).strip(),
                    str(d).strip(),
                    str(correct).strip().upper()
                ])

            QUIZ_CACHE.clear()
            flash(f"✅ Đã upload {len(rows)} câu hỏi từ Excel", "success")
            return redirect(f"/admin?subject={subject}&quiz_id={quiz_id}")

        # ==========================================
        # ===== 4.3 CẬP NHẬT THỜI GIAN LÀM BÀI ======
        # ==========================================
        elif action == "update_time":
            quiz_id = request.form["quiz_id"]
            time_limit = int(request.form["time_limit"])

            rows = ws_list.get_all_values()

            for i in range(1, len(rows)):
                if rows[i][2] == quiz_id:
                    # cột 5 = time_limit
                    if len(rows[i]) < 5:
                        rows[i].append(str(time_limit))
                    else:
                        rows[i][4] = str(time_limit)
                    break

            ws_list.update("A1", rows)
            flash("⏱ Đã cập nhật thời gian làm bài", "success")
            return redirect("/admin")

    # ======================================================
    # ===== 5. RENDER GIAO DIỆN ADMIN ======================
    # ======================================================
    return render_template(
        "admin.html",
        quizzes=quizzes,
        questions=questions,
        selected_quiz=selected_quiz,
        selected_subject=selected_subject
    )




@app.route("/admin/delete-quiz/<subject>/<quiz_id>")
@admin_required
def delete_quiz(subject, quiz_id):
    # xoá trong LIST
    ws_list = sh.worksheet("LIST")
    rows = ws_list.get_all_values()

    for i in range(1, len(rows)):
        if rows[i][2] == quiz_id:
            ws_list.delete_rows(i + 1)
            break

    # xoá trong sheet môn
    sheet_name = SUBJECT_SHEET_MAP.get(subject)
    if sheet_name:
        ws = sh.worksheet(sheet_name)
        all_rows = ws.get_all_values()

        new_rows = [all_rows[0]] + [
            r for r in all_rows[1:] if r[0] != quiz_id
        ]

        ws.clear()
        ws.update("A1", new_rows)

    QUIZ_CACHE.clear()
    flash("🗑️ Đã xoá đề và toàn bộ câu hỏi", "success")
    return redirect("/admin")



# ================= CHỌN MÔN =================
@app.route("/choose-subject")
def choose_subject():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template("choose_subject.html")

# ================= CHỌN LỚP =================
@app.route("/choose-class/<subject>")
def choose_class(subject):
    if "user" not in session:
        return redirect(url_for("login"))

    subject = subject.lower()

    if subject == "trung":
        classes = [
            {"label": f"HSK {i}", "value": f"hsk-{i}"}
            for i in range(1, 7)
        ]

    elif subject in ["ly", "hoa"]:
        classes = [
            {"label": f"Lớp {i}", "value": f"lop-{i}"}
            for i in range(6, 10)
        ]

    else:  # toán
        classes = [
            {"label": f"Lớp {i}", "value": f"lop-{i}"}
            for i in range(1, 10)
        ]

    return render_template(
        "choose_class.html",
        subject=subject,
        classes=classes
    )

# ================= CHỌN ĐỀ =================
@app.route("/list-quiz/<subject>/<class_name>")
def list_quiz(subject, class_name):
    """
    Trang danh sách đề
    subject     : môn học (toan / ly / hoa / trung)
    class_name  : lớp hoặc HSK (dạng URL: hsk-1, lop-3)
    """

    # 1️⃣ Chưa đăng nhập → quay về login
    if "user" not in session:
        return redirect(url_for("login"))

    # 2️⃣ Chuẩn hoá môn học để so sánh
    subject = subject.strip().lower()

    # 3️⃣ Chuẩn hoá lớp:
    #     - URL dùng dấu '-' để tránh lỗi khoảng trắng
    #     - đổi ngược lại để so sánh với Google Sheet
    #     ví dụ: hsk-1  → hsk 1
    #             lop-3 → lop 3
    class_name = class_name.replace("-", " ").strip().lower()
    
    # 4️⃣ Lọc danh sách đề theo *MÔN + LỚP*
    quiz_list = [
        q for q in get_quiz("list")
        if q["subject"].strip().lower() == subject
        and q["class"].strip().lower() == class_name
    ]

    # 5️⃣ Render giao diện danh sách đề
    return render_template(
        "list_quiz.html",
        subject=subject,          # truyền môn sang HTML
        class_name=class_name,    # truyền lớp sang HTML
        quiz_list=quiz_list       # danh sách đề đã lọc
    )



# ================= TRANG LÀM BÀI =================
@app.route("/quiz/<subject>/<quiz_id>")
def quiz(subject, quiz_id):
    if "user" not in session:
        return redirect(url_for("login"))

    quiz_name = ""
    class_name = ""

    for q in get_quiz("list"):
        if q["quiz_id"] == quiz_id:
            quiz_name = q["quiz_name"]
            class_name = q["class"]
            break

    return render_template(
        "quiz.html",
        subject=subject,
        quiz_id=quiz_id,
        quiz_name=quiz_name,
        class_name=class_name
    )




# ================= API LẤY CÂU HỎI =================
@app.route("/api/quiz/<subject>/<quiz_id>")
def api_quiz(subject, quiz_id):
    if "user" not in session:
        return jsonify({"error": "login required"}), 401

    subject = subject.lower()
    quiz_id = str(quiz_id).strip()

    print("SUBJECT:", subject)
    print("QUIZ_ID URL:", quiz_id)
    print("QUIZ_ID SHEET:", [q.get("quiz_id") for q in get_quiz(subject)])

    questions = [
        q for q in get_quiz(subject)
        if str(q.get("quiz_id", "")).strip().upper() == quiz_id.upper()]
    return jsonify(questions)


# ================= SUBMIT + CHẤM ĐIỂM =================
@app.route("/submit/<subject>/<quiz_id>", methods=["POST"])
def submit_quiz(subject, quiz_id):
    print("🔥 REAL SUBMIT HIT 🔥")

    if "user" not in session:
        return jsonify({"error": "login required"}), 401

    data = request.get_json()
    user_answers = data.get("answers", {})

    subject = subject.lower()
    quiz_id = str(quiz_id).strip()

    questions = [
        q for q in get_quiz(subject)
        if str(q.get("quiz_id")).strip().upper() == quiz_id.upper()
    ]

    score = 0
    for idx, q in enumerate(questions):
        if user_answers.get(str(idx)) == q.get("correct_answer"):
            score += 1

    # lấy quiz_name
    quiz_name = ""
    for q in get_quiz("list"):
        if str(q.get("quiz_id")).strip().upper() == quiz_id.upper():
            quiz_name = q.get("quiz_name", "")
            break
    print("STEP 2: before save result")

    # ===== LƯU KẾT QUẢ VÀO SHEET RESULT =====
    try:
        ws_result = sh.worksheet("RESULT")
        ws_result.append_row([
            session.get("user", ""),  # username (đúng với login hiện tại)
            subject,
            quiz_id,
            quiz_name,
            json.dumps(user_answers, ensure_ascii=False),
            score,
            len(questions),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
    except:
        import traceback
        print("SAVE RESULT ERROR:")
        traceback.print_exc()

    print("STEP 3: after save result")

    return jsonify({
        "score": score,
        "total": len(questions)
    })
# ================= XEM KẾT QUẢ =================
@app.route("/result")
def result():
    if "user" not in session:
        return redirect(url_for("login"))

    try:
        ws_result = sh.worksheet("RESULT")
        rows = ws_result.get_all_records()
    except:
        rows = []

    user_rows = [
        r for r in rows
        if r.get("username") == session["user"]
    ]

    return render_template(
        "result.html",
        results=user_rows
    )

# ================= REVIEW BÀI LÀM =================
@app.route("/review/<subject>/<quiz_id>")
def review(subject, quiz_id):
    if "user" not in session:
        return redirect(url_for("login"))

    subject = subject.lower()
    quiz_id = str(quiz_id).strip().upper()

    # 1️⃣ Lấy câu hỏi theo quiz_id
    questions = [
        q for q in get_quiz(subject)
        if str(q.get("quiz_id", "")).strip().upper() == quiz_id
    ]

    # 2️⃣ Lấy kết quả gần nhất của user
    try:
        ws_result = sh.worksheet("RESULT")
        rows = ws_result.get_all_records()
    except:
        rows = []

    user_result = None
    for r in reversed(rows):
        if (
            r.get("username") == session["user"]
            and r.get("subject", "").strip().lower() == subject
            and str(r.get("quiz_id", "")).strip().upper() == quiz_id
        ):
            user_result = r
            break

    # 3️⃣ Parse answers
    user_answers = {}
    quiz_name = ""
    if user_result:
        quiz_name = user_result.get("quiz_name", "")
        try:
            user_answers = json.loads(user_result.get("answers", "{}"))
        except:
            user_answers = {}

    return render_template(
        "review.html",
        subject=subject,
        quiz_id=quiz_id,
        quiz_name=quiz_name,
        questions=questions,
        answers=user_answers
    )




# ================= LOGOUT =================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)